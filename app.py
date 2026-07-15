from fastapi import FastAPI, Request, UploadFile, File, Form
from fastapi.concurrency import run_in_threadpool
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse, RedirectResponse, Response
from pathlib import Path
from dotenv import load_dotenv
load_dotenv()
import csv, io, os, re, sys, uuid, zipfile
from openpyxl import Workbook
from openpyxl.styles import Font
import fitz  # PyMuPDF
from readNestsCSV import parse_nest_csv
from readDrawings import parse_drawing_pdf
import psycopg2
import psycopg2.extras
import psycopg2.pool

BASE_DIR = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / "uploads"

app = FastAPI()

app.mount("/static",   StaticFiles(directory="static"),        name="static")
app.mount("/uploads",  StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")
templates = Jinja2Templates(directory="templates")
templates.env.filters["stem"] = lambda p: Path(p).stem


def _strip_rev(pn: str) -> str:
    return re.sub(r"-R\d+$", "", pn, flags=re.IGNORECASE).upper()


# CSV material name → active materials table column
_MAT_COL = {
    "SPC1.5": "cr4_1_5",        "SPC2.0": "cr4_2_0",
    "SPH3.0": "s275_3_0",       "SPH4.0": "s275_4_0",
    "SPH5.0": "s275_5_0",       "SPH6.0": "s275_6_0",
    "SPH8.0": "s275_8_0",       "SPH10.0": "s275_10_0",
    "SPH5.0-355-GRADE": "s355_5_0", "SPH6.0-355-GRADE": "s355_6_0",
    "SS400-3.0": "s275_3_0",    "SS400-4.0": "s275_4_0",
    "SS400-5.0": "s275_5_0",    "SS400-6.0": "s275_6_0",
    "SS400-8.0": "s275_8_0",    "SS400-10.0": "s275_10_0",
    "SS400-12.0": "s275_12_0",  "SS400-15.0": "s275_15_0",
    "SS400-20.0": "s275_20_0",
    "ALU-5251-H22-1.5MM": "al5251_1_5", "ALU-5251-H22-2MM": "al5251_2_0",
    "ALU-5251-H22-3MM":   "al5251_3_0", "ALU-5251-H22-5MM": "al5251_5_0",
    "A5052-2.0": "al5251_2_0",  "A5052-3.0": "al5251_3_0",
    "A5052-4.0": "al5251_4_0",  "A5052-5.0": "al5251_5_0",
    "GALV-1.0":  "galv_1_0",    "GALV-1.5":  "galv_1_5",
    "GALV-2.0":  "galv_2_0",    "GALV-2.5":  "galv_2_5",
    "GALV-3.0":  "galv_3_0",
    "A1050-1.0": "a1050_1_0",   "A1050-1.5": "a1050_1_5",
    "A1050-2.0": "a1050_2_0",   "A1050-3.0": "a1050_3_0",
    "A1050-4.0": "a1050_4_0",
}

_MAT_LABEL = {
    "SPC1.5": "Mild steel CR4 1.5mm",      "SPC2.0": "Mild steel CR4 2.0mm",
    "SPH4.0": "Hot rolled S275 4.0mm",     "SPH5.0": "Hot rolled S275 5.0mm",
    "SPH6.0": "Hot rolled S275 6.0mm",     "SPH8.0": "Hot rolled S275 8.0mm",
    "SS400-8.0":  "Hot rolled S275 8.0mm", "SS400-10.0": "Hot rolled S275 10.0mm",
    "SS400-12.0": "Hot rolled S275 12.0mm","SS400-15.0": "Hot rolled S275 15.0mm",
    "SS400-20.0": "Hot rolled S275 20.0mm",
    "SPH5.0-355-GRADE": "Hot rolled S355 5.0mm",
    "SPH6.0-355-GRADE": "Hot rolled S355 6.0mm",
    "ALU-5251-H22-2MM": "Aluminium 5251 H22 2.0mm",
    "ALU-5251-H22-3MM": "Aluminium 5251 H22 3.0mm",
    "ALU-5251-H22-5MM": "Aluminium 5251 H22 5.0mm",
    "GALV-1.5": "Galvanised 1.5mm",        "GALV-2.0": "Galvanised 2.0mm",
    "GALV-2.5": "Galvanised 2.5mm",        "GALV-3.0": "Galvanised 3.0mm",
}


# The database is remote (~200ms TCP+TLS+auth per connection), so opening a
# fresh connection per request dominated response times. Connections are
# pooled and health-checked instead; get_db() keeps its old contract —
# callers still call conn.close(), which returns the connection to the pool.

_pool: psycopg2.pool.ThreadedConnectionPool | None = None


def _get_pool() -> psycopg2.pool.ThreadedConnectionPool:
    global _pool
    if _pool is None:
        _pool = psycopg2.pool.ThreadedConnectionPool(
            1, 20, os.getenv("DATABASE_STRING"))
    return _pool


class _PooledConn:
    """Proxy that returns the underlying connection to the pool on close()."""

    def __init__(self, pool, conn):
        self._pool = pool
        self._conn = conn

    def close(self):
        if self._conn is not None:
            try:
                self._conn.rollback()   # clear any open/aborted transaction
                self._pool.putconn(self._conn)
            except psycopg2.Error:
                self._pool.putconn(self._conn, close=True)
            self._conn = None

    def __getattr__(self, name):
        return getattr(self._conn, name)


def get_db():
    try:
        pool = _get_pool()
    except psycopg2.Error:
        return psycopg2.connect(os.getenv("DATABASE_STRING"))

    for _ in range(2):
        try:
            conn = pool.getconn()
        except psycopg2.pool.PoolError:      # pool exhausted
            return psycopg2.connect(os.getenv("DATABASE_STRING"))
        try:
            # Ping: idle pooled connections can be dropped server-side
            cur = conn.cursor()
            cur.execute("SELECT 1")
            cur.close()
            return _PooledConn(pool, conn)
        except psycopg2.Error:
            pool.putconn(conn, close=True)
    return psycopg2.connect(os.getenv("DATABASE_STRING"))


def render_pdf_preview(pdf_path: Path, dpi: int = 150) -> Path | None:
    """Render first page of a PDF to a PNG beside it. Returns the PNG path or None."""
    try:
        png_path = pdf_path.with_suffix(".png")
        doc  = fitz.open(str(pdf_path))
        page = doc[0]
        pix  = page.get_pixmap(matrix=fitz.Matrix(dpi / 72, dpi / 72))
        pix.save(str(png_path))
        doc.close()
        return png_path
    except Exception:
        return None


@app.get("/")
def home(request: Request):
    return templates.TemplateResponse(request, "base.html")


@app.get("/upload")
def upload(request: Request):
    return templates.TemplateResponse(request, "upload.html")


@app.post("/upload")
async def handle_upload(
    csv_file: UploadFile = File(...),
    zip_file: UploadFile = File(...),
    quote_name: str = Form(...),
    customer: str = Form(default=""),
):
    csv_bytes = await csv_file.read()
    zip_bytes = await zip_file.read()
    return await run_in_threadpool(
        _handle_upload_sync, csv_file.filename, csv_bytes, zip_bytes,
        quote_name, customer,
    )


def _handle_upload_sync(csv_filename: str, csv_bytes: bytes, zip_bytes: bytes,
                        quote_name: str, customer: str):
    uid = uuid.uuid4().hex

    # Save CSV
    nest_dir = UPLOAD_DIR / "nests"
    nest_dir.mkdir(parents=True, exist_ok=True)
    csv_path = nest_dir / f"{uid}_{csv_filename}"
    csv_path.write_bytes(csv_bytes)

    # Unzip and save individual PDFs
    drawings_dir = UPLOAD_DIR / "drawings" / uid
    drawings_dir.mkdir(parents=True, exist_ok=True)
    pdf_paths = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for entry in zf.namelist():
            if entry.lower().endswith(".pdf"):
                safe_name = Path(entry).name
                pdf_path = drawings_dir / safe_name
                pdf_path.write_bytes(zf.read(entry))
                render_pdf_preview(pdf_path)
                pdf_paths.append(pdf_path)

    # Store file paths in DB
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO nests (file_path, imported_at) VALUES (%s, NOW()) RETURNING id",
        (csv_path.relative_to(BASE_DIR).as_posix(),),
    )
    nest_id = cur.fetchone()[0]

    cur.execute(
        "INSERT INTO quote_batches (nest_id, name, customer) VALUES (%s, %s, %s) RETURNING id",
        (nest_id, quote_name.strip() or None, customer.strip() or None),
    )
    batch_id = cur.fetchone()[0]

    for pdf_path in pdf_paths:
        cur.execute(
            "INSERT INTO drawings (file_path, imported_at) VALUES (%s, NOW())",
            (pdf_path.relative_to(BASE_DIR).as_posix(),),
        )

    conn.commit()
    cur.close()
    conn.close()

    return JSONResponse({"ok": True, "nest_id": nest_id,
                         "quote_batch_id": batch_id, "pdf_count": len(pdf_paths)})


@app.get("/match/{nest_id}")
def match_review(request: Request, nest_id: int):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cur.execute("SELECT * FROM nests WHERE id = %s", (nest_id,))
    nest = cur.fetchone()
    if not nest:
        cur.close(); conn.close()
        return JSONResponse({"error": "Not found"}, status_code=404)

    uid = Path(nest["file_path"]).name.split("_")[0]
    cur.execute("SELECT * FROM drawings WHERE file_path LIKE %s",
                (f"uploads/drawings/{uid}/%",))
    batch_drawings = list(cur.fetchall())

    # Parts that already have a quote (pricing history)
    cur.execute("""
        SELECT DISTINCT np.part_number
        FROM nest_parts np JOIN quotes q ON q.nest_part_id = np.id
    """)
    priced = {row["part_number"].upper() for row in cur.fetchall()}
    cur.close(); conn.close()

    # Parse each uploaded PDF → build lookup keyed by stripped part number
    drawing_by_part: dict = {}
    for d in batch_drawings:
        # Filename: "LC_TMC_10_BBRP_1 Blade Bolt Reinf Plate.PDF"
        # Split on first space → part number token + description
        stem = Path(d["file_path"]).stem
        stem_parts = stem.split(" ", 1)
        raw_pn_from_name = stem_parts[0].replace("_", "-")   # underscores → hyphens
        desc_from_name   = stem_parts[1] if len(stem_parts) > 1 else ""

        tb, pf, desc = {}, {}, desc_from_name
        key = _strip_rev(raw_pn_from_name)   # filename-based key as default

        try:
            parsed = parse_drawing_pdf(str(BASE_DIR / d["file_path"]))
            if parsed["pages"]:
                tb   = parsed["pages"][0]["title_block"]
                pf   = parsed["pages"][0]["pricing_factors"]
                desc = desc_from_name or tb.get("description") or ""
                tb_pn = tb.get("part_number", "")
                if tb_pn:
                    key = _strip_rev(tb_pn)  # title block is authoritative when present
        except Exception:
            pass

        drawing_by_part[key] = {"db": d, "tb": tb, "pf": pf, "description": desc}

    # Parse CSV
    csv_result = parse_nest_csv(str(BASE_DIR / nest["file_path"]))

    # Match CSV parts → drawings
    matched_parts = []
    used_ids: set = set()
    for part in csv_result["parts"]:
        key = _strip_rev(part["part_number"])
        match = drawing_by_part.get(key)
        if match:
            used_ids.add(match["db"]["id"])
        matched_parts.append({
            **part,
            "drawing":     match["db"] if match else None,
            "description": match["description"] if match else "",
            "has_pricing": key in priced or part["part_number"].upper() in priced,
        })

    unmatched_drawings = [
        {"id": d["id"], "name": Path(d["file_path"]).stem}
        for d in batch_drawings if d["id"] not in used_ids
    ]

    stats = {
        "total":            len(matched_parts),
        "drawings_matched": sum(1 for p in matched_parts if p["drawing"]),
        "no_drawing":       sum(1 for p in matched_parts if not p["drawing"]),
        "not_priced":       sum(1 for p in matched_parts if not p["has_pricing"]),
    }

    return templates.TemplateResponse(request, "match.html", {
        "nest_id":            nest_id,
        "parts":              matched_parts,
        "unmatched_drawings": unmatched_drawings,
        "stats":              stats,
        "csv_warnings":       csv_result.get("_warnings", []),
    })


@app.post("/match/{nest_id}/confirm")
async def confirm_match(request: Request, nest_id: int):
    form = dict(await request.form())
    # PDF re-parsing + DB writes are blocking — keep them off the event loop
    return await run_in_threadpool(_confirm_match_sync, nest_id, form)


def _confirm_match_sync(nest_id: int, form: dict):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM nests WHERE id = %s", (nest_id,))
    nest = cur.fetchone()

    csv_result = parse_nest_csv(str(BASE_DIR / nest["file_path"]))

    wc = conn.cursor()

    # Update nests row with CSV metadata
    wc.execute("""
        UPDATE nests SET source_type = 'csv', schedule_name = %s, total_process_qty = %s
        WHERE id = %s
    """, (csv_result["source_file"], csv_result["schedule"]["total_process_qty"], nest_id))

    for part in csv_result["parts"]:
        pn = part["part_number"]
        drawing_id_str = form.get(f"drawing_{pn}", "")
        drawing_id = int(drawing_id_str) if drawing_id_str and drawing_id_str.strip() else None

        # Update drawings table with parsed PDF metadata
        if drawing_id:
            cur.execute("SELECT file_path FROM drawings WHERE id = %s", (drawing_id,))
            d_row = cur.fetchone()
            if d_row:
                try:
                    parsed = parse_drawing_pdf(str(BASE_DIR / d_row["file_path"]))
                    if parsed["pages"]:
                        tb = parsed["pages"][0]["title_block"]
                        pf = parsed["pages"][0]["pricing_factors"]
                        notes = ", ".join(pf.get("process_notes", []) or []) or None
                        # Filename ("<part> <description>.pdf") is the most
                        # reliable description source — same rule as /match.
                        stem_parts = Path(d_row["file_path"]).stem.split(" ", 1)
                        desc = (stem_parts[1] if len(stem_parts) > 1 else "") \
                               or tb.get("description")
                        wc.execute("""
                            UPDATE drawings SET
                                part_number = %s, revision = %s, description = %s,
                                material = %s, thickness_mm = %s, finish = %s,
                                status = %s, drawn_by = %s,
                                bending_required = %s, bend_count = %s, process_notes = %s
                            WHERE id = %s
                        """, (
                            tb.get("part_number"), tb.get("revision"), desc,
                            tb.get("material"), tb.get("thickness_mm"), tb.get("finish"),
                            tb.get("status"), tb.get("drawn_by"),
                            pf.get("bending_required"), pf.get("bend_count"), notes,
                            drawing_id,
                        ))
                except Exception:
                    pass

        # Upsert so re-confirming the match (back button, refresh) re-imports
        # from the CSV instead of duplicating every part.
        wc.execute("""
            INSERT INTO nest_parts (
                nest_id, drawing_id, part_number,
                size_x_mm, size_y_mm, thickness_mm, material,
                process_time_seconds, total_qty, order_qty
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (nest_id, part_number) DO UPDATE SET
                drawing_id           = EXCLUDED.drawing_id,
                size_x_mm            = EXCLUDED.size_x_mm,
                size_y_mm            = EXCLUDED.size_y_mm,
                thickness_mm         = EXCLUDED.thickness_mm,
                material             = EXCLUDED.material,
                process_time_seconds = EXCLUDED.process_time_seconds,
                total_qty            = EXCLUDED.total_qty,
                order_qty            = EXCLUDED.order_qty
        """, (
            nest_id, drawing_id, pn,
            part.get("size_x_mm"), part.get("size_y_mm"), part.get("thickness_mm"),
            part.get("material"), part.get("process_time_per_part_seconds"),
            part.get("total_qty"), part.get("total_qty"),
        ))

    conn.commit()
    cur.close(); wc.close()

    # Auto-detect assemblies from the uploaded drawings (parts lists / ASSY
    # titles) so the structure page arrives pre-populated. Failures here
    # must never block the pricing flow.
    try:
        _auto_detect_assemblies(conn, nest_id)
    except Exception as e:
        print(f"assembly auto-detect failed for nest {nest_id}: {e}", file=sys.stderr)
    conn.close()

    return RedirectResponse(url=f"/assemblies/{nest_id}", status_code=303)


# ── Assemblies ───────────────────────────────────────────────────────────────
# Parts can be grouped into assemblies, and assemblies nested inside other
# assemblies. Cut/fold is priced on the part; weld / fabrication / machining /
# finishing are priced on the assembly as a whole.

def _assembly_tree(assemblies: list[dict]) -> list[dict]:
    """Annotate each assembly with depth and total builds (qty × parent
    builds), returning a depth-first flattened list (parents before children).
    Orphans from a broken parent chain are appended at depth 0."""
    by_id     = {a["id"]: a for a in assemblies}
    by_parent: dict = {}
    for a in assemblies:
        pid = a["parent_assembly_id"]
        by_parent.setdefault(pid if pid in by_id else None, []).append(a)

    out: list[dict] = []

    def walk(pid, depth, mult):
        for a in by_parent.get(pid, []):
            q = max(a.get("qty") or 1, 1)
            a["depth"]  = depth
            a["builds"] = mult * q
            out.append(a)
            walk(a["id"], depth + 1, mult * q)

    walk(None, 0, 1)
    seen = {a["id"] for a in out}
    for a in assemblies:   # cycle fallback — never reached with validated parents
        if a["id"] not in seen:
            a["depth"], a["builds"] = 0, max(a.get("qty") or 1, 1)
            out.append(a)
    return out


def fetch_assemblies(cur, nest_id: int) -> list[dict]:
    """All assemblies for a nest as a flattened tree with builds/depth/part count."""
    cur.execute("""
        SELECT a.*, COUNT(np.id) AS part_count
        FROM assemblies a
        LEFT JOIN nest_parts np ON np.assembly_id = a.id
        WHERE a.nest_id = %s
        GROUP BY a.id
        ORDER BY a.sort_order, a.id
    """, (nest_id,))
    return _assembly_tree([dict(r) for r in cur.fetchall()])


def _is_descendant(cur, assembly_id: int, candidate_parent: int) -> bool:
    """True if candidate_parent sits below assembly_id (or is it) — would cycle."""
    cur.execute("SELECT id, parent_assembly_id FROM assemblies WHERE nest_id = "
                "(SELECT nest_id FROM assemblies WHERE id = %s)", (assembly_id,))
    rows = cur.fetchall()
    # Works with both tuple and RealDict cursors
    parent_of = {(r["id"] if isinstance(r, dict) else r[0]):
                 (r["parent_assembly_id"] if isinstance(r, dict) else r[1])
                 for r in rows}
    node = candidate_parent
    while node is not None:
        if node == assembly_id:
            return True
        node = parent_of.get(node)
    return False


_ASSY_NAME_RE = re.compile(r"\b(ASSY|ASSEMBLY|WELDMENT)\b", re.I)


def _norm_pn(pn: str) -> str:
    """Normalise a part number for matching: underscores → hyphens, strip
    trailing revision, uppercase — same rules as the drawing/CSV matcher."""
    return _strip_rev((pn or "").replace("_", "-").strip())


def _auto_detect_assemblies(conn, nest_id: int) -> dict:
    """Scan the nest's uploaded drawings for assembly drawings (parts list /
    BOM tables, ASSY/WELDMENT titles) and build the assembly tree from them:
    create an assembly per assembly drawing, assign nest parts referenced in
    its BOM, and nest sub-assemblies whose BOM rows point at other assembly
    drawings (with the BOM quantity).

    Parse results are cached on the drawings row (is_assembly, bom), so
    re-running only parses drawings that were never scanned. Manual structure
    is respected: parts already assigned and assemblies already nested are
    never moved."""
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM nests WHERE id = %s", (nest_id,))
    nest = cur.fetchone()
    if not nest:
        cur.close()
        return {"error": "Nest not found"}

    uid = Path(nest["file_path"]).name.split("_")[0]
    cur.execute("SELECT * FROM drawings WHERE file_path LIKE %s",
                (f"uploads/drawings/{uid}/%",))
    batch_drawings = [dict(r) for r in cur.fetchall()]

    wc = conn.cursor()
    infos, scanned = [], 0
    for d in batch_drawings:
        stem       = Path(d["file_path"]).stem
        stem_parts = stem.split(" ", 1)
        info = {
            "db":   d,
            "key":  _norm_pn(stem_parts[0]),
            "name": (stem_parts[1] if len(stem_parts) > 1 else stem).strip(),
            "asm":  None,
        }

        if d.get("is_assembly") is not None:
            # Cached from an earlier scan — no re-parse
            if d["is_assembly"]:
                info["asm"] = {"is_assembly": True, "bom": d.get("bom") or []}
            if d.get("part_number"):
                info["key"] = _norm_pn(d["part_number"])
        else:
            scanned += 1
            asm = None
            try:
                parsed = parse_drawing_pdf(str(BASE_DIR / d["file_path"]))
                page0  = parsed["pages"][0] if parsed["pages"] else {}
                tb     = page0.get("title_block") or {}
                asm    = page0.get("assembly")
                if tb.get("part_number"):
                    info["key"] = _norm_pn(tb["part_number"])
                if tb.get("description") and len(stem_parts) == 1:
                    info["name"] = tb["description"]
            except Exception:
                pass
            info["asm"] = asm
            wc.execute("UPDATE drawings SET is_assembly = %s, bom = %s WHERE id = %s",
                       (bool(asm),
                        psycopg2.extras.Json(asm["bom"]) if asm else None,
                        d["id"]))

        # Filename that says ASSY/WELDMENT counts even with no parts list
        if not info["asm"] and _ASSY_NAME_RE.search(stem):
            info["asm"] = {"is_assembly": True, "bom": []}
        infos.append(info)

    assembly_infos = [i for i in infos if i["asm"]]

    # Existing assemblies → idempotent re-runs (match by drawing, then name)
    cur.execute("SELECT * FROM assemblies WHERE nest_id = %s", (nest_id,))
    existing   = [dict(r) for r in cur.fetchall()]
    by_drawing = {a["drawing_id"]: a for a in existing if a.get("drawing_id")}
    by_name    = {a["name"].strip().upper(): a for a in existing}

    created = 0
    info_by_key = {}
    for i in assembly_infos:
        a = by_drawing.get(i["db"]["id"]) or by_name.get(i["name"].upper())
        if a:
            i["assembly_id"] = a["id"]
        else:
            wc.execute("""
                INSERT INTO assemblies (nest_id, name, qty, drawing_id, auto_detected)
                VALUES (%s, %s, 1, %s, TRUE) RETURNING id
            """, (nest_id, i["name"][:120], i["db"]["id"]))
            i["assembly_id"] = wc.fetchone()[0]
            created += 1
        info_by_key[i["key"]] = i

    # Resolve BOM rows: nest parts get assigned, assembly drawings get nested
    cur.execute("SELECT id, part_number, assembly_id FROM nest_parts WHERE nest_id = %s",
                (nest_id,))
    part_by_key = {_norm_pn(r["part_number"]): dict(r) for r in cur.fetchall()}

    assigned = linked = 0
    for i in assembly_infos:
        for row in i["asm"]["bom"] or []:
            key = _norm_pn(row.get("part_number", ""))
            if not key:
                continue
            child = info_by_key.get(key)
            if child and child["assembly_id"] != i["assembly_id"]:
                # Sub-assembly: nest it (never re-parent something the user
                # placed, and never create a cycle)
                if not _is_descendant(cur, child["assembly_id"], i["assembly_id"]):
                    wc.execute("""
                        UPDATE assemblies SET parent_assembly_id = %s, qty = %s
                        WHERE id = %s AND parent_assembly_id IS NULL
                    """, (i["assembly_id"], max(int(row.get("qty") or 1), 1),
                          child["assembly_id"]))
                    linked += wc.rowcount
            else:
                p = part_by_key.get(key)
                if p and p["assembly_id"] is None:
                    wc.execute("""
                        UPDATE nest_parts SET assembly_id = %s
                        WHERE id = %s AND assembly_id IS NULL
                    """, (i["assembly_id"], p["id"]))
                    assigned += wc.rowcount
                    p["assembly_id"] = i["assembly_id"]

    conn.commit()
    cur.close(); wc.close()
    return {"ok": True, "drawings_scanned": scanned,
            "assemblies_detected": len(assembly_infos),
            "assemblies_created": created,
            "parts_assigned": assigned, "sub_assemblies_linked": linked}


@app.post("/assemblies/{nest_id}/autodetect")
def assemblies_autodetect(nest_id: int):
    """Re-run assembly auto-detection over this nest's drawings."""
    conn = get_db()
    try:
        result = _auto_detect_assemblies(conn, nest_id)
    finally:
        conn.close()
    status = 404 if result.get("error") else 200
    return JSONResponse(result, status_code=status)


@app.get("/assemblies/{nest_id}")
def assemblies_page(request: Request, nest_id: int):
    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT np.id, np.part_number, np.material, np.total_qty, np.assembly_id,
               d.description
        FROM nest_parts np LEFT JOIN drawings d ON d.id = np.drawing_id
        WHERE np.nest_id = %s ORDER BY np.id
    """, (nest_id,))
    parts = [dict(r) for r in cur.fetchall()]
    assemblies = fetch_assemblies(cur, nest_id)
    cur.close(); conn.close()
    return templates.TemplateResponse(request, "assemblies.html", {
        "nest_id": nest_id, "parts": parts, "assemblies": assemblies,
    })


@app.post("/assemblies")
async def assembly_create(request: Request):
    data = await request.json()
    name = (data.get("name") or "").strip()
    if not name:
        return JSONResponse({"error": "Name required"}, status_code=400)
    parent = data.get("parent_assembly_id") or None
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("""
        INSERT INTO assemblies (nest_id, parent_assembly_id, name, qty)
        VALUES (%s, %s, %s, %s) RETURNING id
    """, (data["nest_id"], parent, name, max(int(data.get("qty") or 1), 1)))
    new_id = cur.fetchone()[0]
    conn.commit()
    cur.close(); conn.close()
    return JSONResponse({"ok": True, "id": new_id})


@app.post("/assemblies/{assembly_id}/update")
async def assembly_update(request: Request, assembly_id: int):
    data = await request.json()
    conn = get_db()
    cur  = conn.cursor()

    if "parent_assembly_id" in data:
        parent = data["parent_assembly_id"] or None
        if parent and _is_descendant(cur, assembly_id, int(parent)):
            cur.close(); conn.close()
            return JSONResponse({"error": "That would nest an assembly inside itself"},
                                status_code=400)
        cur.execute("UPDATE assemblies SET parent_assembly_id = %s WHERE id = %s",
                    (parent, assembly_id))

    if data.get("name", "").strip():
        cur.execute("UPDATE assemblies SET name = %s WHERE id = %s",
                    (data["name"].strip(), assembly_id))
    if data.get("qty") is not None:
        cur.execute("UPDATE assemblies SET qty = %s WHERE id = %s",
                    (max(int(data["qty"]), 1), assembly_id))

    conn.commit()
    cur.close(); conn.close()
    return JSONResponse({"ok": True})


@app.post("/assemblies/{assembly_id}/delete")
def assembly_delete(assembly_id: int):
    conn = get_db()
    cur  = conn.cursor()
    # Promote children to the deleted assembly's parent; parts unassign via FK
    cur.execute("""
        UPDATE assemblies SET parent_assembly_id =
            (SELECT parent_assembly_id FROM assemblies WHERE id = %s)
        WHERE parent_assembly_id = %s
    """, (assembly_id, assembly_id))
    cur.execute("DELETE FROM assembly_quotes WHERE assembly_id = %s", (assembly_id,))
    cur.execute("DELETE FROM assemblies WHERE id = %s", (assembly_id,))
    deleted = cur.rowcount
    conn.commit()
    cur.close(); conn.close()
    if not deleted:
        return JSONResponse({"error": "Not found"}, status_code=404)
    return JSONResponse({"ok": True})


@app.post("/assemblies/{nest_id}/assign")
async def assembly_assign(request: Request, nest_id: int):
    """Assign a nest part to an assembly (or back to loose with null)."""
    data        = await request.json()
    assembly_id = data.get("assembly_id") or None
    conn = get_db()
    cur  = conn.cursor()
    if assembly_id:
        cur.execute("SELECT 1 FROM assemblies WHERE id = %s AND nest_id = %s",
                    (assembly_id, nest_id))
        if not cur.fetchone():
            cur.close(); conn.close()
            return JSONResponse({"error": "Assembly not in this nest"}, status_code=400)
    cur.execute("UPDATE nest_parts SET assembly_id = %s WHERE id = %s AND nest_id = %s",
                (assembly_id, data["nest_part_id"], nest_id))
    conn.commit()
    cur.close(); conn.close()
    return JSONResponse({"ok": True})


# ── Pricing calculation ──────────────────────────────────────────────────────

def calculate_price(d: dict, rates: dict) -> dict:
    """Compute part pricing from form inputs + a burden_rates row.

    Single source of truth for the pricing math — used by the live-calc
    endpoint and by the save flow, so stored totals always match the UI.
    """
    def f(k, default=0):
        try: return float(d.get(k) or default)
        except (TypeError, ValueError): return float(default)
    def iv(k, default=0):
        try: return int(d.get(k) or default)
        except (TypeError, ValueError): return int(default)

    qty    = max(iv("qty", 1), 1)
    active = d.get("active_processes") or []

    laser  = ((f("size_x_mm") * f("size_y_mm") / 1_000_000)
               * f("material_cost_m2") * rates["scrap_factor"] * qty
               + (rates["laser_hourly_rate"] / 60) * (f("process_time_seconds") / 60) * qty)

    fold   = 0.0
    nf     = iv("num_folds")
    if nf > 0:
        mf   = f("mins_per_fold", 0.25)
        fs   = f("fold_setup_mins", 5)
        fold = (qty * nf * mf + nf * mf + fs) * (rates["fold_rate"] / 60)

    tube = (f("tube_cut_time_mins") * (rates["tube_hourly_rate"] / 60) * qty
            + f("tube_kg_per_metre") * f("tube_length_m") * f("tube_cost_per_kg")
            * rates["scrap_factor"] * qty) if "tube" in active else 0.0

    weld     = (f("weld_time_mins") / 60) * rates["weld_saw_rate"] * qty if "weld" in active else 0.0
    saw      = ((f("saw_setup_mins") + iv("saw_num_cuts") * f("saw_mins_per_cut"))
                * (rates["weld_saw_rate"] / 60)) if "saw" in active else 0.0
    machine  = (f("machine_time_mins")  / 60) * rates["machine_rate"]  * qty if "machine"  in active else 0.0
    assembly = (f("assembly_time_mins") / 60) * rates["assembly_rate"] * qty if "assembly" in active else 0.0

    finish   = f("finish_cost") * rates["finishing_markup"] * qty
    purchase = f("purchased")   * qty
    sticker  = f("sticker_cost",  0.07) * qty
    delivery = f("delivery_cost", 5.0)  * qty
    misc     = f("misc_cost")           * qty

    total      = laser + fold + tube + weld + saw + machine + assembly + finish + purchase + sticker + delivery + misc
    margin_pct = f("margin_pct", 10)

    return {
        "cost_per_part": round(total / qty, 2),
        "total_cost":    round(total, 2),
        "margin_total":  round(total * (1 + margin_pct / 100), 2),
        "breakdown": {
            "laser": round(laser, 2), "fold":     round(fold, 2),
            "tube":  round(tube, 2),  "weld":     round(weld, 2),
            "saw":   round(saw, 2),   "machine":  round(machine, 2),
            "assembly": round(assembly, 2),
            "finish":   round(finish, 2),  "purchased": round(purchase, 2),
            "sticker":  round(sticker, 2), "delivery":  round(delivery, 2),
            "misc":     round(misc, 2),
        },
    }


def calculate_assembly_price(d: dict, rates: dict) -> dict:
    """Compute assembly-level process costs (weld / fabrication / machining /
    finishing) from form inputs + a burden_rates row. Inputs are per build;
    'builds' is the total number of this assembly across the order
    (qty × parent builds). Child part/assembly costs are rolled up separately."""
    def f(k, default=0):
        try: return float(d.get(k) or default)
        except (TypeError, ValueError): return float(default)

    builds = max(int(f("builds", 1)), 1)
    active = d.get("active_processes") or []

    weld    = (f("weld_time_mins")    / 60) * rates["weld_saw_rate"] * builds if "weld"    in active else 0.0
    fab     = (f("fab_time_mins")     / 60) * rates["assembly_rate"] * builds if "fab"     in active else 0.0
    machine = (f("machine_time_mins") / 60) * rates["machine_rate"]  * builds if "machine" in active else 0.0
    finish  = f("finish_cost") * rates["finishing_markup"] * builds           if "finish"  in active else 0.0
    misc    = f("misc_cost") * builds

    own        = weld + fab + machine + finish + misc
    margin_pct = f("margin_pct", 10)

    return {
        "builds":           builds,
        "cost_per_build":   round(own / builds, 2),
        "own_cost":         round(own, 2),
        "own_margin_total": round(own * (1 + margin_pct / 100), 2),
        "breakdown": {
            "weld":    round(weld, 2),    "fab":    round(fab, 2),
            "machine": round(machine, 2), "finish": round(finish, 2),
            "misc":    round(misc, 2),
        },
    }


def assembly_quote_row_to_state(q: dict) -> dict:
    """Map a saved assembly_quotes row back to the pricing-form state shape."""
    active = [proc for proc in ("weld", "fab", "machine", "finish")
              if q.get(f"{proc}_active")]
    return {
        "assembly_id":       q.get("assembly_id"),
        "builds":            q.get("builds"),
        "weld_time_mins":    q.get("weld_time_mins"),
        "fab_time_mins":     q.get("fab_time_mins"),
        "machine_time_mins": q.get("machine_time_mins"),
        "finish_cost":       q.get("finish_cost"),
        "misc_cost":         q.get("misc_cost"),
        "margin_pct":        q.get("margin_pct"),
        "active_processes":  active,
        "own_cost":          q.get("own_cost"),
        "own_margin_total":  (round(q["own_cost"] * (1 + (q.get("margin_pct") or 0) / 100), 2)
                              if q.get("own_cost") is not None else None),
    }


def quote_row_to_state(q: dict) -> dict:
    """Map a saved quotes row back to the pricing-form state shape used by the
    front end (sessionStorage / save payload), so quotes can be reopened."""
    active = []
    if (q.get("num_folds") or 0) > 0:
        active.append("fold")
    # Weld / machine / assembly are assembly-level processes now — old part
    # quotes that had them are re-priced with cut/fold only.
    for proc in ("tube", "saw"):
        if q.get(f"{proc}_active"):
            active.append(proc)
    return {
        "nest_part_id":       q.get("nest_part_id"),
        "qty":                q.get("quantity"),
        "material_cost_m2":   q.get("material_cost_m2"),
        "num_folds":          q.get("num_folds"),
        "mins_per_fold":      q.get("mins_per_fold"),
        "fold_setup_mins":    q.get("fold_setup_mins"),
        "purchased":          q.get("purchased_total"),
        "finish_cost":        q.get("finish_cost_per_part"),
        "sticker_cost":       q.get("sticker_cost"),
        "delivery_cost":      q.get("delivery_cost"),
        "misc_cost":          q.get("misc_cost"),
        "margin_pct":         q.get("margin_pct"),
        "active_processes":   active,
        "tube_cut_time_mins": q.get("tube_cut_time_mins"),
        "tube_kg_per_metre":  q.get("tube_kg_per_metre"),
        "tube_length_m":      q.get("tube_length_m"),
        "tube_cost_per_kg":   q.get("tube_cost_per_kg"),
        "weld_time_mins":     q.get("weld_time_mins"),
        "saw_num_cuts":       q.get("saw_num_cuts"),
        "saw_mins_per_cut":   q.get("saw_mins_per_cut"),
        "saw_setup_mins":     q.get("saw_setup_mins"),
        "machine_time_mins":  q.get("machine_time_mins"),
        "assembly_time_mins": q.get("assembly_time_mins"),
        "cost_per_part":      q.get("cost_per_part"),
        "total_cost":         q.get("line_cost"),
        "margin_total":       q.get("margin_total"),
        "complete":           q.get("material_cost_m2") is not None,
    }


# ── Pricing pages ────────────────────────────────────────────────────────────

@app.get("/price/{nest_id}")
def price_start(nest_id: int):
    return RedirectResponse(url=f"/price/{nest_id}/1", status_code=302)


@app.get("/price/{nest_id}/review")
def price_review(request: Request, nest_id: int):
    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT np.*, d.description AS drawing_desc
        FROM nest_parts np
        LEFT JOIN drawings d ON d.id = np.drawing_id
        WHERE np.nest_id = %s ORDER BY np.id
    """, (nest_id,))
    parts = [dict(p) for p in cur.fetchall()]

    # Saved lines from the latest batch — fallback for parts the browser has
    # no sessionStorage state for, so editing a saved quote keeps its values.
    cur.execute("SELECT id FROM quote_batches WHERE nest_id = %s ORDER BY id DESC LIMIT 1",
                (nest_id,))
    batch = cur.fetchone()
    saved = {}
    if batch:
        cur.execute("SELECT * FROM quotes WHERE quote_batch_id = %s", (batch["id"],))
        saved = {r["nest_part_id"]: quote_row_to_state(dict(r)) for r in cur.fetchall()}

    cur.execute("SELECT * FROM burden_rates WHERE is_active = true LIMIT 1")
    rates = cur.fetchone()
    cur.execute("SELECT * FROM materials WHERE is_active = true LIMIT 1")
    mats  = cur.fetchone()

    # Pricing history for parts with no saved line, fetched in ONE query
    # (latest quote per part number) — the remote DB makes per-part queries
    # cost ~20ms each, which added seconds on large nests.
    need_history = [p["part_number"] for p in parts if p["id"] not in saved]
    history: dict = {}
    if need_history and rates:
        cur.execute("""
            SELECT DISTINCT ON (UPPER(np2.part_number))
                   UPPER(np2.part_number) AS hist_pn, q.*
            FROM quotes q
            JOIN nest_parts np2 ON np2.id = q.nest_part_id
            WHERE UPPER(np2.part_number) = ANY(%s)
            ORDER BY UPPER(np2.part_number),
                     q.quoted_at DESC NULLS LAST, q.id DESC
        """, ([pn.upper() for pn in need_history],))
        history = {r["hist_pn"]: dict(r) for r in cur.fetchall()}

    for p in parts:
        state = saved.get(p["id"])
        if not state and rates:
            # Same pricing-history fallback as the part page, recomputed with
            # this nest's quantity/geometry and current rates — otherwise
            # history-prefilled parts show as Incomplete until each is opened.
            hist = history.get(p["part_number"].upper())
            if hist:
                state = quote_row_to_state(dict(hist))
                state["nest_part_id"] = p["id"]
                state["qty"]          = p["total_qty"]
                col = _MAT_COL.get(p["material"])
                if col and mats and mats.get(col) is not None:
                    state["material_cost_m2"] = mats[col]
                calc = calculate_price({
                    **state,
                    "size_x_mm":            p["size_x_mm"],
                    "size_y_mm":            p["size_y_mm"],
                    "process_time_seconds": p["process_time_seconds"],
                }, rates)
                state.update(
                    size_x_mm=p["size_x_mm"], size_y_mm=p["size_y_mm"],
                    process_time_seconds=p["process_time_seconds"],
                    cost_per_part=calc["cost_per_part"],
                    total_cost=calc["total_cost"],
                    margin_total=calc["margin_total"],
                    complete=bool(state.get("material_cost_m2")),
                    prefill_source="history",
                )
        p["saved"] = state

    # Assembly tree + saved assembly-level pricing (latest batch)
    assemblies  = fetch_assemblies(cur, nest_id)
    saved_assy = {}
    if batch:
        cur.execute("SELECT * FROM assembly_quotes WHERE quote_batch_id = %s",
                    (batch["id"],))
        saved_assy = {r["assembly_id"]: assembly_quote_row_to_state(dict(r))
                      for r in cur.fetchall()}
    for a in assemblies:
        state = saved_assy.get(a["id"])
        if state:
            state["builds"] = a["builds"]   # structure may have changed since save
        a["saved"] = state
        a.pop("created_at", None)           # not JSON-serialisable, not needed

    cur.close(); conn.close()
    return templates.TemplateResponse(request, "review.html", {
        "nest_id": nest_id, "parts": parts, "assemblies": assemblies,
    })


def _opt_num(v):
    """Parse an optional numeric form value; '' / None / junk → None."""
    try:
        return float(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def _opt_text(v):
    v = (v or "").strip()
    return v or None


@app.post("/price/{nest_id}/save")
async def price_save(request: Request, nest_id: int):
    data = await request.json()
    # Long chain of blocking queries — keep it off the event loop
    return await run_in_threadpool(_price_save_sync, nest_id, data)


def _price_save_sync(nest_id: int, data):
    # Old payload shape was a bare list of parts; new shape splits
    # part-level and assembly-level pricing.
    if isinstance(data, list):
        parts_data, assy_data = data, []
    else:
        parts_data = data.get("parts") or []
        assy_data  = data.get("assemblies") or []
    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cur.execute("SELECT * FROM burden_rates WHERE is_active = true LIMIT 1")
    rates = cur.fetchone()
    if not rates:
        cur.close(); conn.close()
        return JSONResponse({"error": "No active burden rates"}, status_code=400)

    cur.execute("SELECT id FROM materials WHERE is_active = true LIMIT 1")
    mat_row     = cur.fetchone()
    material_id = mat_row["id"] if mat_row else None

    # Latest batch for this nest (created at upload; legacy nests get one now)
    cur.execute("SELECT id FROM quote_batches WHERE nest_id = %s ORDER BY id DESC LIMIT 1",
                (nest_id,))
    batch = cur.fetchone()
    if batch:
        batch_id = batch["id"]
    else:
        cur.execute("INSERT INTO quote_batches (nest_id) VALUES (%s) RETURNING id", (nest_id,))
        batch_id = cur.fetchone()["id"]

    cur.execute("SELECT id, drawing_id, assembly_id FROM nest_parts WHERE nest_id = %s",
                (nest_id,))
    rows        = cur.fetchall()
    drawing_of  = {r["id"]: r["drawing_id"]  for r in rows}
    assembly_of = {r["id"]: r["assembly_id"] for r in rows}

    wc = conn.cursor()
    subtotal = margin_sum = 0.0
    part_results: dict = {}
    for p in parts_data:
        np_id = p.get("nest_part_id")
        if np_id not in drawing_of:
            continue

        result = calculate_price(p, rates)
        part_results[np_id] = result
        subtotal   += result["total_cost"]
        margin_sum += result["margin_total"]

        qty = _opt_num(p.get("qty"))

        # Part-detail edits from the pricing page back onto the part / drawing
        wc.execute("""
            UPDATE nest_parts SET
                size_x_mm            = COALESCE(%s, size_x_mm),
                size_y_mm            = COALESCE(%s, size_y_mm),
                thickness_mm         = COALESCE(%s, thickness_mm),
                material             = COALESCE(%s, material),
                process_time_seconds = COALESCE(%s, process_time_seconds),
                order_qty            = COALESCE(%s, order_qty)
            WHERE id = %s
        """, (
            _opt_num(p.get("size_x_mm")), _opt_num(p.get("size_y_mm")),
            _opt_num(p.get("ph_thk")),    _opt_text(p.get("ph_mat")),
            _opt_num(p.get("process_time_seconds")), qty, np_id,
        ))
        if drawing_of[np_id]:
            wc.execute("""
                UPDATE drawings SET
                    revision    = COALESCE(%s, revision),
                    description = COALESCE(%s, description)
                WHERE id = %s
            """, (_opt_text(p.get("ph_rev")), _opt_text(p.get("ph_desc")), drawing_of[np_id]))

        active = p.get("active_processes") or []
        wc.execute("""
            INSERT INTO quotes (
                quote_batch_id, nest_part_id, burden_rate_id, material_id,
                quoted_at, quantity,
                purchased_total, finish_cost_per_part, sticker_cost, delivery_cost,
                misc_cost, margin_pct,
                material_cost_m2, num_folds, mins_per_fold, fold_setup_mins,
                tube_active, weld_active, saw_active, machine_active, assembly_active,
                tube_cut_time_mins, tube_kg_per_metre, tube_length_m, tube_cost_per_kg,
                weld_time_mins, saw_num_cuts, saw_mins_per_cut, saw_setup_mins,
                machine_time_mins, assembly_time_mins,
                cost_per_part, line_cost, margin_total, breakdown
            ) VALUES (
                %s,%s,%s,%s, NOW(),%s, %s,%s,%s,%s,%s,%s,
                %s,%s,%s,%s,
                %s,%s,%s,%s,%s,
                %s,%s,%s,%s, %s, %s,%s,%s, %s,%s,
                %s,%s,%s,%s
            )
            ON CONFLICT (quote_batch_id, nest_part_id) WHERE quote_batch_id IS NOT NULL
            DO UPDATE SET
                burden_rate_id = EXCLUDED.burden_rate_id,
                material_id    = EXCLUDED.material_id,
                quoted_at      = EXCLUDED.quoted_at,
                quantity       = EXCLUDED.quantity,
                purchased_total      = EXCLUDED.purchased_total,
                finish_cost_per_part = EXCLUDED.finish_cost_per_part,
                sticker_cost   = EXCLUDED.sticker_cost,
                delivery_cost  = EXCLUDED.delivery_cost,
                misc_cost      = EXCLUDED.misc_cost,
                margin_pct     = EXCLUDED.margin_pct,
                material_cost_m2 = EXCLUDED.material_cost_m2,
                num_folds        = EXCLUDED.num_folds,
                mins_per_fold    = EXCLUDED.mins_per_fold,
                fold_setup_mins  = EXCLUDED.fold_setup_mins,
                tube_active      = EXCLUDED.tube_active,
                weld_active      = EXCLUDED.weld_active,
                saw_active       = EXCLUDED.saw_active,
                machine_active   = EXCLUDED.machine_active,
                assembly_active  = EXCLUDED.assembly_active,
                tube_cut_time_mins = EXCLUDED.tube_cut_time_mins,
                tube_kg_per_metre  = EXCLUDED.tube_kg_per_metre,
                tube_length_m      = EXCLUDED.tube_length_m,
                tube_cost_per_kg   = EXCLUDED.tube_cost_per_kg,
                weld_time_mins     = EXCLUDED.weld_time_mins,
                saw_num_cuts       = EXCLUDED.saw_num_cuts,
                saw_mins_per_cut   = EXCLUDED.saw_mins_per_cut,
                saw_setup_mins     = EXCLUDED.saw_setup_mins,
                machine_time_mins  = EXCLUDED.machine_time_mins,
                assembly_time_mins = EXCLUDED.assembly_time_mins,
                cost_per_part = EXCLUDED.cost_per_part,
                line_cost     = EXCLUDED.line_cost,
                margin_total  = EXCLUDED.margin_total,
                breakdown     = EXCLUDED.breakdown
        """, (
            batch_id, np_id, rates["id"], material_id,
            qty or 1,
            p.get("purchased", 0), p.get("finish_cost", 0),
            p.get("sticker_cost", 0.07), p.get("delivery_cost", 5),
            p.get("misc_cost", 0),       p.get("margin_pct", 10),
            _opt_num(p.get("material_cost_m2")), _opt_num(p.get("num_folds")),
            _opt_num(p.get("mins_per_fold")),    _opt_num(p.get("fold_setup_mins")),
            "tube" in active, "weld" in active, "saw" in active,
            "machine" in active, "assembly" in active,
            p.get("tube_cut_time_mins"), p.get("tube_kg_per_metre"),
            p.get("tube_length_m"),      p.get("tube_cost_per_kg"),
            p.get("weld_time_mins"),
            p.get("saw_num_cuts"),       p.get("saw_mins_per_cut"),
            p.get("saw_setup_mins"),     p.get("machine_time_mins"),
            p.get("assembly_time_mins"),
            result["cost_per_part"], result["total_cost"], result["margin_total"],
            psycopg2.extras.Json(result["breakdown"]),
        ))

    # ── Assembly-level costs + tree roll-up ──
    assemblies = fetch_assemblies(cur, nest_id)
    assy_state = {a["assembly_id"]: a for a in assy_data if a.get("assembly_id")}

    # Child part line costs / margins per assembly
    part_cost_of: dict = {}
    part_marg_of: dict = {}
    for np_id, res in part_results.items():
        aid = assembly_of.get(np_id)
        if aid:
            part_cost_of[aid] = part_cost_of.get(aid, 0.0) + res["total_cost"]
            part_marg_of[aid] = part_marg_of.get(aid, 0.0) + res["margin_total"]

    children_of: dict = {}
    for a in assemblies:
        if a["parent_assembly_id"]:
            children_of.setdefault(a["parent_assembly_id"], []).append(a["id"])

    own_results = {a["id"]: calculate_assembly_price(
        {**(assy_state.get(a["id"]) or {}), "builds": a["builds"]}, rates,
    ) for a in assemblies}

    # assemblies is flattened parent-first, so reversed() walks children first
    rolled_cost: dict = {}
    rolled_marg: dict = {}
    for a in reversed(assemblies):
        aid = a["id"]
        own = own_results[aid]
        rolled_cost[aid] = round(own["own_cost"] + part_cost_of.get(aid, 0.0)
                                 + sum(rolled_cost[c] for c in children_of.get(aid, [])), 2)
        rolled_marg[aid] = round(own["own_margin_total"] + part_marg_of.get(aid, 0.0)
                                 + sum(rolled_marg[c] for c in children_of.get(aid, [])), 2)

    for a in assemblies:
        aid    = a["id"]
        state  = assy_state.get(aid) or {}
        own    = own_results[aid]
        active = state.get("active_processes") or []
        subtotal   += own["own_cost"]
        margin_sum += own["own_margin_total"]
        wc.execute("""
            INSERT INTO assembly_quotes (
                quote_batch_id, assembly_id, quoted_at, builds,
                weld_active, weld_time_mins, fab_active, fab_time_mins,
                machine_active, machine_time_mins, finish_active, finish_cost,
                misc_cost, margin_pct,
                own_cost, rolled_cost, rolled_margin_total, breakdown
            ) VALUES (%s,%s, NOW(),%s, %s,%s,%s,%s, %s,%s,%s,%s, %s,%s, %s,%s,%s,%s)
            ON CONFLICT (quote_batch_id, assembly_id) DO UPDATE SET
                quoted_at           = EXCLUDED.quoted_at,
                builds              = EXCLUDED.builds,
                weld_active         = EXCLUDED.weld_active,
                weld_time_mins      = EXCLUDED.weld_time_mins,
                fab_active          = EXCLUDED.fab_active,
                fab_time_mins       = EXCLUDED.fab_time_mins,
                machine_active      = EXCLUDED.machine_active,
                machine_time_mins   = EXCLUDED.machine_time_mins,
                finish_active       = EXCLUDED.finish_active,
                finish_cost         = EXCLUDED.finish_cost,
                misc_cost           = EXCLUDED.misc_cost,
                margin_pct          = EXCLUDED.margin_pct,
                own_cost            = EXCLUDED.own_cost,
                rolled_cost         = EXCLUDED.rolled_cost,
                rolled_margin_total = EXCLUDED.rolled_margin_total,
                breakdown           = EXCLUDED.breakdown
        """, (
            batch_id, aid, a["builds"],
            "weld" in active,    _opt_num(state.get("weld_time_mins")),
            "fab" in active,     _opt_num(state.get("fab_time_mins")),
            "machine" in active, _opt_num(state.get("machine_time_mins")),
            "finish" in active,  _opt_num(state.get("finish_cost")),
            _opt_num(state.get("misc_cost")), state.get("margin_pct", 10),
            own["own_cost"], rolled_cost[aid], rolled_marg[aid],
            psycopg2.extras.Json(own["breakdown"]),
        ))

    wc.execute("""
        UPDATE quote_batches SET
            status = 'final', burden_rate_id = %s, material_id = %s,
            subtotal = %s, total_with_margin = %s, finalized_at = NOW()
        WHERE id = %s
    """, (rates["id"], material_id, round(subtotal, 2), round(margin_sum, 2), batch_id))

    conn.commit()
    cur.close(); wc.close(); conn.close()
    return JSONResponse({"ok": True, "quote_batch_id": batch_id})


@app.post("/price/calculate")
async def price_calculate(request: Request):
    d    = await request.json()
    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM burden_rates WHERE is_active = true LIMIT 1")
    rates = cur.fetchone()
    cur.close(); conn.close()
    if not rates:
        return JSONResponse({"error": "No active burden rates"}, status_code=400)

    return JSONResponse(calculate_price(d, rates))


@app.post("/price/assembly/calculate")
async def price_assembly_calculate(request: Request):
    d    = await request.json()
    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM burden_rates WHERE is_active = true LIMIT 1")
    rates = cur.fetchone()
    cur.close(); conn.close()
    if not rates:
        return JSONResponse({"error": "No active burden rates"}, status_code=400)

    return JSONResponse(calculate_assembly_price(d, rates))


@app.get("/price/{nest_id}/assembly/{assembly_id}")
def price_assembly(request: Request, nest_id: int, assembly_id: int):
    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    assemblies = fetch_assemblies(cur, nest_id)
    assembly   = next((a for a in assemblies if a["id"] == assembly_id), None)
    if not assembly:
        cur.close(); conn.close()
        return JSONResponse({"error": "Not found"}, status_code=404)

    # Contents: direct child parts and sub-assemblies
    cur.execute("""
        SELECT np.id, np.part_number, np.total_qty, d.description
        FROM nest_parts np LEFT JOIN drawings d ON d.id = np.drawing_id
        WHERE np.nest_id = %s AND np.assembly_id = %s ORDER BY np.id
    """, (nest_id, assembly_id))
    child_parts = [dict(r) for r in cur.fetchall()]
    child_assemblies = [a for a in assemblies
                        if a["parent_assembly_id"] == assembly_id]

    # Saved assembly pricing (latest batch) → prefill fallback
    cur.execute("""
        SELECT aq.* FROM assembly_quotes aq
        JOIN quote_batches qb ON qb.id = aq.quote_batch_id
        WHERE aq.assembly_id = %s AND qb.nest_id = %s
        ORDER BY qb.id DESC LIMIT 1
    """, (assembly_id, nest_id))
    saved_row   = cur.fetchone()
    saved_state = assembly_quote_row_to_state(dict(saved_row)) if saved_row else None
    if saved_state:
        saved_state["builds"] = assembly["builds"]
    cur.close(); conn.close()

    # Prev/next assembly for the nav bar (tree order)
    idx = next(i for i, a in enumerate(assemblies) if a["id"] == assembly_id)
    return templates.TemplateResponse(request, "price_assembly.html", {
        "nest_id":          nest_id,
        "assembly":         assembly,
        "assemblies":       assemblies,
        "assembly_index":   idx + 1,
        "prev_id":          assemblies[idx - 1]["id"] if idx > 0 else None,
        "next_id":          assemblies[idx + 1]["id"] if idx + 1 < len(assemblies) else None,
        "child_parts":      child_parts,
        "child_assemblies": child_assemblies,
        "saved_state":      saved_state,
    })


@app.get("/price/{nest_id}/{part_index}")
def price_part(request: Request, nest_id: int, part_index: int):
    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT np.*, d.file_path AS drawing_path, d.description AS drawing_desc,
               d.revision AS drawing_revision, d.bending_required, d.bend_count,
               a.name AS assembly_name
        FROM nest_parts np
        LEFT JOIN drawings d ON d.id = np.drawing_id
        LEFT JOIN assemblies a ON a.id = np.assembly_id
        WHERE np.nest_id = %s ORDER BY np.id
    """, (nest_id,))
    all_parts = [dict(p) for p in cur.fetchall()]

    if not all_parts or part_index < 1 or part_index > len(all_parts):
        cur.close(); conn.close()
        return JSONResponse({"error": "Not found"}, status_code=404)

    part = all_parts[part_index - 1]

    col          = _MAT_COL.get(part["material"])
    material_cost = None
    if col:
        cur.execute(f"SELECT {col} FROM materials WHERE is_active = true LIMIT 1")
        row = cur.fetchone()
        if row:
            material_cost = row[col]

    # Previously saved line for this part (latest batch) → prefill when the
    # browser has no sessionStorage state, so saved quotes can be edited.
    cur.execute("""
        SELECT q.* FROM quotes q
        JOIN quote_batches qb ON qb.id = q.quote_batch_id
        WHERE q.nest_part_id = %s AND qb.nest_id = %s
        ORDER BY qb.id DESC LIMIT 1
    """, (part["id"], nest_id))
    saved_row   = cur.fetchone()
    saved_state = None
    if saved_row:
        saved_state = quote_row_to_state(dict(saved_row))
    else:
        # Pricing history: the same part number quoted on any earlier nest.
        cur.execute("""
            SELECT q.* FROM quotes q
            JOIN nest_parts np ON np.id = q.nest_part_id
            WHERE UPPER(np.part_number) = UPPER(%s)
            ORDER BY q.quoted_at DESC NULLS LAST, q.id DESC
            LIMIT 1
        """, (part["part_number"],))
        hist = cur.fetchone()
        if hist:
            saved_state = quote_row_to_state(dict(hist))
            saved_state["nest_part_id"] = part["id"]
            # Job-specific values come from the current nest, not history
            for k in ("qty", "cost_per_part", "total_cost", "margin_total", "complete"):
                saved_state.pop(k, None)
            if material_cost is not None:
                # Current price list beats the historic material cost
                saved_state.pop("material_cost_m2", None)
            saved_state["prefill_source"] = "history"
            saved_state["prefill_date"]   = (hist["quoted_at"].strftime("%d %b %Y")
                                             if hist["quoted_at"] else None)

    assemblies = fetch_assemblies(cur, nest_id)
    cur.close(); conn.close()

    drawing_image_url = None
    if part.get("drawing_path"):
        png_rel = part["drawing_path"].rsplit(".", 1)[0] + ".png"
        if (BASE_DIR / png_rel).exists():
            drawing_image_url = "/" + png_rel

    return templates.TemplateResponse(request, "price.html", {
        "nest_id":           nest_id,
        "part_index":        part_index,
        "total_parts":       len(all_parts),
        "part":              part,
        "part_numbers":      [p["part_number"] for p in all_parts],
        "material_cost":     material_cost,
        "catalogue":         _MAT_LABEL.get(part["material"], ""),
        "drawing_image_url": drawing_image_url,
        "saved_state":       saved_state,
        "first_assembly_id": assemblies[0]["id"] if assemblies else None,
    })


@app.get("/rates")
def rates(request: Request):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM materials ORDER BY is_active DESC NULLS LAST, created_at DESC;")
    mat_rows = cur.fetchall()
    cur.execute("SELECT * FROM burden_rates ORDER BY is_active DESC NULLS LAST, created_at DESC;")
    rates_rows = cur.fetchall()
    cur.close()
    conn.close()
    return templates.TemplateResponse(request, "index.html", {"materials": mat_rows, "burden_rates": rates_rows})


@app.post("/materials/{material_id}/select")
def select_material(material_id: int):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE materials SET is_active = false")
    cur.execute("UPDATE materials SET is_active = true WHERE id = %s", (material_id,))
    conn.commit()
    cur.close()
    conn.close()
    return JSONResponse({"ok": True})


@app.post("/materials")
async def add_material_set(request: Request):
    data = await request.json()
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO materials (
          name,
          cr4_1_5, cr4_2_0,
          s275_3_0, s275_4_0, s275_5_0, s275_6_0, s275_8_0, s275_10_0, s275_12_0, s275_15_0, s275_20_0,
          s355_3_0, s355_4_0, s355_5_0, s355_6_0, s355_8_0, s355_10_0, s355_12_0, s355_15_0, s355_20_0,
          hardox450_4_0, hardox450_5_0, hardox450_6_0, hardox450_8_0, hardox450_10_0, hardox450_12_0, hardox450_15_0, hardox450_20_0,
          ss304dp1_0_9, ss304dp1_1_5, ss304dp1_2_0, ss304dp1_3_0, ss304dp1_4_0, ss304dp1_5_0, ss304dp1_6_0, ss304dp1_8_0, ss304dp1_10_0,
          al5251_1_0, al5251_2_0, al5251_3_0, al5251_4_0, al5251_5_0, al5251_6_0, al5251_8_0, al5251_10_0,
          galv_1_0, galv_1_5, galv_2_0, galv_2_5, galv_3_0,
          a1050_1_0, a1050_1_5, a1050_2_0, a1050_3_0, a1050_4_0,
          created_at, is_active
        ) VALUES (
          %s,
          %s,%s,
          %s,%s,%s,%s,%s,%s,%s,%s,%s,
          %s,%s,%s,%s,%s,%s,%s,%s,%s,
          %s,%s,%s,%s,%s,%s,%s,%s,
          %s,%s,%s,%s,%s,%s,%s,%s,%s,
          %s,%s,%s,%s,%s,%s,%s,%s,
          %s,%s,%s,%s,%s,
          %s,%s,%s,%s,%s,
          NOW(), false
        )
    """, (
        data["name"],
        data.get("cr4_1_5"), data.get("cr4_2_0"),
        data.get("s275_3_0"), data.get("s275_4_0"), data.get("s275_5_0"), data.get("s275_6_0"), data.get("s275_8_0"), data.get("s275_10_0"), data.get("s275_12_0"), data.get("s275_15_0"), data.get("s275_20_0"),
        data.get("s355_3_0"), data.get("s355_4_0"), data.get("s355_5_0"), data.get("s355_6_0"), data.get("s355_8_0"), data.get("s355_10_0"), data.get("s355_12_0"), data.get("s355_15_0"), data.get("s355_20_0"),
        data.get("hardox450_4_0"), data.get("hardox450_5_0"), data.get("hardox450_6_0"), data.get("hardox450_8_0"), data.get("hardox450_10_0"), data.get("hardox450_12_0"), data.get("hardox450_15_0"), data.get("hardox450_20_0"),
        data.get("ss304dp1_0_9"), data.get("ss304dp1_1_5"), data.get("ss304dp1_2_0"), data.get("ss304dp1_3_0"), data.get("ss304dp1_4_0"), data.get("ss304dp1_5_0"), data.get("ss304dp1_6_0"), data.get("ss304dp1_8_0"), data.get("ss304dp1_10_0"),
        data.get("al5251_1_0"), data.get("al5251_2_0"), data.get("al5251_3_0"), data.get("al5251_4_0"), data.get("al5251_5_0"), data.get("al5251_6_0"), data.get("al5251_8_0"), data.get("al5251_10_0"),
        data.get("galv_1_0"), data.get("galv_1_5"), data.get("galv_2_0"), data.get("galv_2_5"), data.get("galv_3_0"),
        data.get("a1050_1_0"), data.get("a1050_1_5"), data.get("a1050_2_0"), data.get("a1050_3_0"), data.get("a1050_4_0"),
    ))
    conn.commit()
    cur.close()
    conn.close()
    return JSONResponse({"ok": True})


@app.post("/rates/{rate_id}/select")
def select_rate(rate_id: int):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE burden_rates SET is_active = false")
    cur.execute("UPDATE burden_rates SET is_active = true WHERE id = %s", (rate_id,))
    conn.commit()
    cur.close()
    conn.close()
    return JSONResponse({"ok": True})


# ── Saved quotes & export ────────────────────────────────────────────────────

def fetch_quote_batch(batch_id: int):
    """Load a quote batch header + its part lines + its assembly lines
    (tree-ordered, with depth)."""
    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT qb.*, br.name AS rate_name
        FROM quote_batches qb
        LEFT JOIN burden_rates br ON br.id = qb.burden_rate_id
        WHERE qb.id = %s
    """, (batch_id,))
    batch = cur.fetchone()
    lines, assembly_lines = [], []
    if batch:
        cur.execute("SELECT COUNT(*) AS n FROM nest_parts WHERE nest_id = %s",
                    (batch["nest_id"],))
        batch["part_count"] = cur.fetchone()["n"]
        cur.execute("""
            SELECT q.*, np.part_number, np.material, np.thickness_mm,
                   np.assembly_id, d.revision, d.description
            FROM quotes q
            JOIN nest_parts np ON np.id = q.nest_part_id
            LEFT JOIN drawings d ON d.id = np.drawing_id
            WHERE q.quote_batch_id = %s
            ORDER BY q.id
        """, (batch_id,))
        lines = [dict(r) for r in cur.fetchall()]
        for l in lines:
            # Customer-facing price each (margin included)
            qty = l["quantity"] or 1
            l["price_each"] = round((l["margin_total"] or 0) / qty, 2)

        cur.execute("""
            SELECT a.id, a.name, a.parent_assembly_id, a.qty, a.sort_order,
                   aq.builds AS saved_builds, aq.margin_pct, aq.own_cost,
                   aq.rolled_cost, aq.rolled_margin_total, aq.breakdown,
                   aq.weld_time_mins, aq.fab_time_mins, aq.machine_time_mins,
                   aq.finish_cost, aq.misc_cost
            FROM assembly_quotes aq
            JOIN assemblies a ON a.id = aq.assembly_id
            WHERE aq.quote_batch_id = %s
            ORDER BY a.sort_order, a.id
        """, (batch_id,))
        assembly_lines = _assembly_tree([dict(r) for r in cur.fetchall()])
        for a in assembly_lines:
            # Show the builds count the quote was saved with, if any
            a["builds"] = a["saved_builds"] or a["builds"]
            a["price_each"] = round((a["rolled_margin_total"] or 0)
                                    / max(a["builds"], 1), 2)
    cur.close(); conn.close()
    return batch, lines, assembly_lines


def quote_display_rows(lines, assembly_lines):
    """Interleave assembly header rows with their part lines, depth-first;
    parts in no (or an unsaved) assembly follow at the end at depth 0."""
    parts_of: dict = {}
    for l in lines:
        parts_of.setdefault(l.get("assembly_id"), []).append(l)

    rows, shown = [], set()
    for a in assembly_lines:   # already tree-ordered with depth
        a["n_parts"] = len(parts_of.get(a["id"], []))
        a["n_subs"]  = sum(1 for o in assembly_lines
                           if o["parent_assembly_id"] == a["id"])
        rows.append({"type": "assembly", "depth": a["depth"], "a": a})
        for l in parts_of.get(a["id"], []):
            rows.append({"type": "part", "depth": a["depth"] + 1, "l": l})
            shown.add(l["id"])
    for l in lines:
        if l["id"] not in shown:
            rows.append({"type": "part", "depth": 0, "l": l})
    return rows


def _quote_filename(batch) -> str:
    # JMS convention: "QUO - 031815 - Groundsman"
    if batch["quote_no"]:
        name = f"QUO - {batch['quote_no']} - {batch['customer'] or batch['name'] or batch['id']}"
    else:
        name = batch["name"] or f"quote_{batch['id']}"
    return re.sub(r"[^A-Za-z0-9 _-]+", "_", name).strip("_ ") or f"quote_{batch['id']}"


@app.post("/quotes/{batch_id}/meta")
async def quote_meta(request: Request, batch_id: int):
    data = await request.json()
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("""
        UPDATE quote_batches SET
            quote_no  = %s, contact = %s, raised_by = %s, notes = %s,
            name      = COALESCE(%s, name),
            customer  = COALESCE(%s, customer)
        WHERE id = %s
    """, (
        _opt_text(data.get("quote_no")), _opt_text(data.get("contact")),
        _opt_text(data.get("raised_by")), _opt_text(data.get("notes")),
        _opt_text(data.get("name")), _opt_text(data.get("customer")),
        batch_id,
    ))
    updated = cur.rowcount
    conn.commit()
    cur.close(); conn.close()
    if not updated:
        return JSONResponse({"error": "Not found"}, status_code=404)
    return JSONResponse({"ok": True})


@app.get("/quotes")
def quotes_list(request: Request):
    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT qb.*, COUNT(q.id) AS line_count,
               (SELECT COUNT(*) FROM nest_parts np WHERE np.nest_id = qb.nest_id) AS part_count
        FROM quote_batches qb
        LEFT JOIN quotes q ON q.quote_batch_id = qb.id
        GROUP BY qb.id
        ORDER BY qb.created_at DESC
    """)
    batches = cur.fetchall()
    cur.close(); conn.close()
    return templates.TemplateResponse(request, "quotes.html", {"batches": batches})


@app.post("/quotes/{batch_id}/delete")
def quote_delete(batch_id: int):
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("DELETE FROM quotes WHERE quote_batch_id = %s", (batch_id,))
    cur.execute("DELETE FROM assembly_quotes WHERE quote_batch_id = %s", (batch_id,))
    cur.execute("DELETE FROM quote_batches WHERE id = %s", (batch_id,))
    deleted = cur.rowcount
    conn.commit()
    cur.close(); conn.close()
    if not deleted:
        return JSONResponse({"error": "Not found"}, status_code=404)
    return JSONResponse({"ok": True})


@app.get("/quotes/{batch_id}")
def quote_detail(request: Request, batch_id: int):
    batch, lines, assembly_lines = fetch_quote_batch(batch_id)
    if not batch:
        return JSONResponse({"error": "Not found"}, status_code=404)
    return templates.TemplateResponse(request, "quote.html", {
        "batch": batch, "lines": lines,
        "rows":  quote_display_rows(lines, assembly_lines),
    })


_EXPORT_HEADERS = ["Part No", "Rev", "Description", "Material", "Thk (mm)", "Qty",
                   "Cost / Part", "Line Cost", "With Margin", "Margin %"]


def _export_rows(rows):
    """Flatten display rows (assemblies + parts) for CSV export, indenting by depth."""
    for r in rows:
        pad = "  " * r["depth"]
        if r["type"] == "assembly":
            a = r["a"]
            yield [pad + "[ASSY] " + a["name"], "", "Assembly - weld/fab/machine/finish",
                   "", "", a["builds"],
                   round((a["rolled_cost"] or 0) / max(a["builds"], 1), 2),
                   a["rolled_cost"], a["rolled_margin_total"], a["margin_pct"]]
        else:
            l = r["l"]
            yield [pad + l["part_number"], l["revision"] or "", l["description"] or "",
                   l["material"] or "", l["thickness_mm"], l["quantity"],
                   l["cost_per_part"], l["line_cost"], l["margin_total"], l["margin_pct"]]


@app.get("/quotes/{batch_id}/export.csv")
def quote_export_csv(batch_id: int):
    batch, lines, assembly_lines = fetch_quote_batch(batch_id)
    if not batch:
        return JSONResponse({"error": "Not found"}, status_code=404)
    rows = quote_display_rows(lines, assembly_lines)

    buf = io.StringIO()
    w   = csv.writer(buf)
    quoted = batch["finalized_at"] or batch["created_at"]
    w.writerow(["Quotation No.", batch["quote_no"] or batch["id"]])
    w.writerow(["Quote",         batch["name"] or f"Quote {batch['id']}"])
    w.writerow(["Customer",      batch["customer"] or ""])
    w.writerow(["Contact",       batch["contact"] or ""])
    w.writerow(["Raised by",     batch["raised_by"] or ""])
    w.writerow(["Date",          quoted.strftime("%Y-%m-%d %H:%M") if quoted else ""])
    w.writerow(["Rate set",      batch["rate_name"] or ""])
    w.writerow([])
    w.writerow(_EXPORT_HEADERS)
    for row in _export_rows(rows):
        w.writerow(row)
    w.writerow([])
    w.writerow(["Totals", "", "", "", "", "", "", batch["subtotal"], batch["total_with_margin"], ""])

    return Response(buf.getvalue(), media_type="text/csv", headers={
        "Content-Disposition": f'attachment; filename="{_quote_filename(batch)}.csv"',
    })


# breakdown key → JMS Summary-sheet process row (same order as their workbook).
# "fab" is assembly-level fabrication; "assembly"/"weld"/"machine"/"finish" also
# appear in legacy part-level breakdowns and in assembly-level breakdowns.
_JMS_PROCESS_ROWS = [
    ("Laser cut",   ("laser",)),
    ("Tube cut",    ("tube",)),
    ("Brake press", ("fold",)),
    ("Saw cut",     ("saw",)),
    ("Machine",     ("machine",)),
    ("Welding",     ("weld",)),
    ("Assembly",    ("assembly", "fab")),
    ("Finishing",   ("finish",)),
    ("Misc",        ("misc", "sticker", "delivery", "purchased")),
]

_GBP_FMT = '"£"#,##0.00'

# materials-table column prefix → grade label (for the Material Prices sheet)
_MAT_GRADE_LABELS = {
    "cr4":       "Mild steel CR4",
    "s275":      "Hot rolled S275",
    "s355":      "Hot rolled S355",
    "hardox450": "Hardox 450",
    "ss304dp1":  "Stainless 304 DP1",
    "al5251":    "Aluminium 5251",
    "galv":      "Galvanised",
    "a1050":     "Aluminium 1050",
}


def _material_price_rows(mat_row: dict):
    """Yield (grade label, thickness mm, £/m²) from a materials table row."""
    for col, val in mat_row.items():
        if col in ("id", "name", "created_at", "is_active") or val is None:
            continue
        tokens = col.split("_")
        if len(tokens) < 3:
            continue
        prefix = "_".join(tokens[:-2])
        try:
            thickness = float(f"{tokens[-2]}.{tokens[-1]}")
        except ValueError:
            continue
        yield _MAT_GRADE_LABELS.get(prefix, prefix), thickness, val


@app.get("/quotes/{batch_id}/export.xlsx")
def quote_export_xlsx(batch_id: int):
    batch, lines, assembly_lines = fetch_quote_batch(batch_id)
    if not batch:
        return JSONResponse({"error": "Not found"}, status_code=404)
    display_rows = quote_display_rows(lines, assembly_lines)

    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM burden_rates WHERE id = %s", (batch["burden_rate_id"],))
    rates = cur.fetchone()
    if batch["material_id"]:
        cur.execute("SELECT * FROM materials WHERE id = %s", (batch["material_id"],))
    else:
        cur.execute("SELECT * FROM materials WHERE is_active = true LIMIT 1")
    mat_row = cur.fetchone()
    cur.close(); conn.close()

    bold   = Font(bold=True)
    wb     = Workbook()
    quoted = batch["finalized_at"] or batch["created_at"]

    # ── Summary sheet — mirrors the JMS internal costing workbook layout ──
    ws = wb.active
    ws.title = "Summary"
    ws["B1"] = "Customer";        ws["C1"] = batch["customer"] or ""
    ws["B2"] = "ENQ / SO number"; ws["C2"] = (f"ENQ - {batch['quote_no']}"
                                              if batch["quote_no"] else batch["name"] or "")
    ws["B3"] = "Date submitted";  ws["C3"] = quoted
    ws["C3"].number_format = "d-mmm-yy"
    for c in ("B1", "B2", "B3"):
        ws[c].font = bold

    ws["B5"] = "Process"; ws["C5"] = "£"
    ws["B5"].font = ws["C5"].font = bold

    totals = {}
    for item in lines + assembly_lines:
        for key, val in (item["breakdown"] or {}).items():
            totals[key] = totals.get(key, 0.0) + (val or 0)

    r = 6
    for label, keys in _JMS_PROCESS_ROWS:
        ws.cell(row=r, column=2, value=label)
        cell = ws.cell(row=r, column=3, value=round(sum(totals.get(k, 0) for k in keys), 2))
        cell.number_format = _GBP_FMT
        r += 1

    ws.cell(row=r, column=2, value="Total").font = bold
    c_total = ws.cell(row=r, column=3, value=batch["subtotal"])
    c_total.number_format = _GBP_FMT; c_total.font = bold
    c_marg = ws.cell(row=r, column=4, value=batch["total_with_margin"])
    c_marg.number_format = _GBP_FMT; c_marg.font = bold
    ws.cell(row=r + 1, column=4, value="(with margin)")

    # Hourly-rates panel on the right, like "2026 Hourly Rates"
    if rates:
        ws["G5"] = f"Hourly Rates — {rates['name'] or ''}".strip()
        ws["G5"].font = bold
        rate_rows = [
            ("Flat Laser",       rates["laser_hourly_rate"]),
            ("Tube Laser",       rates["tube_hourly_rate"]),
            ("Brake Presses",    rates["fold_rate"]),
            ("Welding / Saw",    rates["weld_saw_rate"]),
            ("Machine",          rates["machine_rate"]),
            ("Assembly",         rates["assembly_rate"]),
            ("Finishing markup", rates["finishing_markup"]),
            ("Scrap factor",     rates["scrap_factor"]),
        ]
        for i, (label, val) in enumerate(rate_rows, start=6):
            ws.cell(row=i, column=7, value=label)
            cell = ws.cell(row=i, column=8, value=val)
            if label not in ("Finishing markup", "Scrap factor"):
                cell.number_format = _GBP_FMT

    for col, width in (("B", 28.9), ("C", 14), ("D", 14), ("G", 19.9), ("H", 10.7)):
        ws.column_dimensions[col].width = width

    # ── Parts sheet — customer-facing line items (assembly tree order) ──
    # Assembly rows carry the rolled-up subtotal (own processes + contents);
    # the indented part rows beneath show the make-up, so only top-level rows
    # (depth 0) sum to the NET total.
    ps = wb.create_sheet("Parts")
    ps.append(["Line", "Part Ref", "Description", "Material", "Thk (mm)",
               "Rev No.", "Qty.", "Price ea", "Subtotal"])
    for cell in ps[1]:
        cell.font = bold
    for i, r in enumerate(display_rows, start=1):
        pad = "    " * r["depth"]
        if r["type"] == "assembly":
            a = r["a"]
            ps.append([i, pad + "⌞ " + a["name"] if r["depth"] else pad + a["name"],
                       "Assembly", "", "", "", a["builds"],
                       a["price_each"], a["rolled_margin_total"]])
            ps.cell(row=ps.max_row, column=2).font = bold
        else:
            l = r["l"]
            ps.append([i, pad + l["part_number"], l["description"] or "",
                       l["material"] or "", l["thickness_mm"], l["revision"] or "",
                       l["quantity"], l["price_each"], l["margin_total"]])
        ps.cell(row=ps.max_row, column=8).number_format = _GBP_FMT
        ps.cell(row=ps.max_row, column=9).number_format = _GBP_FMT
    ps.append([])
    ps.append(["", "", "", "", "", "", "NET Total", "", batch["total_with_margin"]])
    ps.cell(row=ps.max_row, column=7).font = bold
    net = ps.cell(row=ps.max_row, column=9)
    net.number_format = _GBP_FMT; net.font = bold
    for col, width in zip("ABCDEFGHI", (6, 18, 36, 12, 9, 9, 7, 11, 11)):
        ps.column_dimensions[col].width = width

    # ── Breakdown sheet — per-part per-process costs ──
    procs = ["laser", "fold", "tube", "weld", "saw", "machine", "assembly",
             "finish", "purchased", "sticker", "delivery", "misc"]
    bs = wb.create_sheet("Breakdown")
    bs.append(["Part No", "Material £/m²"] + [p.capitalize() for p in procs] + ["Line Cost"])
    for cell in bs[1]:
        cell.font = bold
    for l in lines:
        bd = l["breakdown"] or {}
        bs.append([l["part_number"], l["material_cost_m2"]]
                  + [bd.get(p) for p in procs] + [l["line_cost"]])

    # Assembly-level process costs (weld / fabrication / machine / finish)
    if assembly_lines:
        bs.append([])
        bs.append(["Assembly", "Builds", "Weld", "Fabrication", "Machine",
                   "Finish", "Misc", "Own Cost", "Rolled Cost"])
        for cell in bs[bs.max_row]:
            cell.font = bold
        for a in assembly_lines:
            bd = a["breakdown"] or {}
            bs.append(["  " * a["depth"] + a["name"], a["builds"],
                       bd.get("weld"), bd.get("fab"), bd.get("machine"),
                       bd.get("finish"), bd.get("misc"),
                       a["own_cost"], a["rolled_cost"]])
    bs.column_dimensions["A"].width = 16
    bs.column_dimensions["B"].width = 13

    # ── Material Prices sheet — the price set this quote was costed with ──
    if mat_row:
        ms = wb.create_sheet("Material Prices")
        ms.append([f"Material Prices — {mat_row['name'] or ''}".strip()])
        ms["A1"].font = bold
        ms.append(["Grade", "Thickness (mm)", "£/m²"])
        for cell in ms[2]:
            cell.font = bold
        for grade, thickness, price in _material_price_rows(dict(mat_row)):
            ms.append([grade, thickness, price])
            ms.cell(row=ms.max_row, column=3).number_format = _GBP_FMT
        for col, width in (("A", 20), ("B", 14), ("C", 10)):
            ms.column_dimensions[col].width = width

    stream = io.BytesIO()
    wb.save(stream)
    return Response(
        stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{_quote_filename(batch)}.xlsx"'},
    )


@app.post("/rates")
async def add_rate(request: Request):
    data = await request.json()
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """INSERT INTO burden_rates
           (name, laser_hourly_rate, tube_hourly_rate, fold_rate, weld_saw_rate,
            machine_rate, assembly_rate, finishing_markup, scrap_factor,
            mins_per_fold, fold_setup_mins, created_at, is_active)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s, NOW(), false)""",
        (data["name"], float(data["laser_hourly_rate"]), float(data["tube_hourly_rate"]),
         float(data["fold_rate"]), float(data["weld_saw_rate"]), float(data["machine_rate"]),
         float(data["assembly_rate"]), float(data["finishing_markup"]), float(data["scrap_factor"]),
         float(data["mins_per_fold"]), float(data["fold_setup_mins"]))
    )
    conn.commit()
    cur.close()
    conn.close()
    return JSONResponse({"ok": True})
